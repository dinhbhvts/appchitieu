"""One-off cleanup: remove StockTrade rows created by a bug in
extract_stock_rows() (now fixed in import_excel.py).

The bug: the buy/sell scan did not stop at the "ĐANG GIỮ" (currently-held)
snapshot below the trade log, and that snapshot reuses the same columns
(D=symbol, E=quantity, F=price) as a buy order - so every holding line got
imported as an extra "buy" trade. Example: for T7.2026 this created 15
phantom trades (9 for Vợ, 6 for Chồng) on top of the 1 real trade actually
recorded in the sheet.

This script re-scans each sheet's "ĐANG GIỮ" block the same way the OLD
buggy code used to read it (mirroring its buy/sell detection exactly), and
for every phantom entry found, deletes ONE matching StockTrade row (same
date/symbol/side/quantity/price/owner) from the database. It never deletes
more rows than phantom entries found, so:
  - A trade you entered by hand in the app is untouched (it has no matching
    phantom entry to "consume" it).
  - If a real trade happens to have the exact same numbers as a holding
    (e.g. you just bought GEL and it now also shows in ĐANG GIỮ with the
    same qty/price), only ONE of the two identical-looking rows is removed -
    the correct trade count is restored either way, since the two rows are
    indistinguishable in the app anyway.

Dry run by default (prints what WOULD be deleted). Add --commit to delete.

Run (from the backend folder, virtual environment active, DATABASE_URL set
to the real database in .env):
    python -m scripts.cleanup_phantom_stock_trades
    python -m scripts.cleanup_phantom_stock_trades --commit
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.core.database import SessionLocal  # noqa: E402
from app.models.enums import TradeSide  # noqa: E402
from app.models.stock import StockTrade  # noqa: E402
from app.models.user import User  # noqa: E402
from scripts.import_excel import (  # noqa: E402
    SKIP_SHEETS,
    _is_symbol,
    _stock_owner,
    _strip_accents,
    parse_month_year,
    resolve_day,
    safe_date,
)

_STOCK_SKIP = ("total", "tong", "loi nhuan", "lai/lo", "lai lo")


def find_phantom_trades(rows: list, year: int, month: int) -> list[dict]:
    """Reproduce exactly what the OLD buggy extract_stock_rows() used to
    wrongly generate from the "ĐANG GIỮ" block for one sheet."""
    sub = None
    for i, raw in enumerate(rows):
        r = list(raw) + [None] * 20
        if _strip_accents(r[3]) == "ma ck":
            sub = i
            break
    if sub is None:
        return []

    label_idx = None
    last_day = 1
    for i in range(sub + 1, len(rows)):
        r = list(rows[i]) + [None] * 20
        if _strip_accents(r[3]) == "dang giu":
            label_idx = i
            break
        # Track last_day up to the label, same as the old loop would have.
        last_day = resolve_day(r[0], year, month, last_day)
    if label_idx is None:
        return []

    phantom: list[dict] = []
    for i in range(label_idx, len(rows)):
        r = list(rows[i]) + [None] * 20
        note_text = r[1]
        b = _strip_accents(note_text)
        # The holdings block itself ends at its own TOTAL/summary line -
        # bound the scan there (matches extract_holdings()'s own boundary).
        if any(k in b for k in _STOCK_SKIP):
            break
        last_day = resolve_day(r[0], year, month, last_day)
        d = safe_date(year, month, last_day)
        owner = _stock_owner(r, b)
        fee = float(r[11]) if isinstance(r[11], (int, float)) else 0.0

        if _is_symbol(r[3]) and isinstance(r[4], (int, float)) and r[4] > 0 \
                and isinstance(r[5], (int, float)) and r[5] > 0:
            phantom.append({"date": d, "symbol": r[3].strip().upper(),
                             "side": TradeSide.buy, "quantity": int(r[4]),
                             "price": float(r[5]), "fee": fee, "owner": owner})
        if _is_symbol(r[7]) and isinstance(r[8], (int, float)) and r[8] > 0 \
                and isinstance(r[9], (int, float)) and r[9] > 0:
            phantom.append({"date": d, "symbol": r[7].strip().upper(),
                             "side": TradeSide.sell, "quantity": int(r[8]),
                             "price": float(r[9]), "fee": fee, "owner": owner})
    return phantom


def run(excel_path: Path, commit: bool) -> None:
    with SessionLocal() as db:
        users = {u.name: u.id for u in db.query(User).all()}

        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

        total_found = 0
        total_deleted = 0
        by_month: list[str] = []

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            my = parse_month_year(sheet_name)
            if my is None:
                continue
            year, month = my
            rows = list(wb[sheet_name].iter_rows(values_only=True))

            phantom = find_phantom_trades(rows, year, month)
            if not phantom:
                continue

            deleted_here = 0
            for p in phantom:
                uid = users.get(p["owner"])
                if uid is None:
                    continue
                row = (
                    db.query(StockTrade)
                    .filter(
                        StockTrade.date == p["date"],
                        StockTrade.symbol == p["symbol"],
                        StockTrade.side == p["side"],
                        StockTrade.quantity == p["quantity"],
                        StockTrade.price == p["price"],
                        StockTrade.user_id == uid,
                    )
                    .first()
                )
                if row is not None:
                    total_found += 1
                    deleted_here += 1
                    if commit:
                        db.delete(row)
                        total_deleted += 1

            if deleted_here:
                by_month.append(
                    f"  {sheet_name}: {deleted_here} dong phantom "
                    f"({'da xoa' if commit else 'se xoa neu --commit'})"
                )

        if commit:
            db.commit()

        print(f"Tong so dong phantom tim thay: {total_found}")
        for line in by_month:
            print(line)
        if commit:
            print(f"\nDa xoa {total_deleted} dong khoi stock_trades.")
        else:
            print("\n(Dry run - chua xoa gi. Chay lai voi --commit de xoa that.)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    default_path = Path(__file__).resolve().parents[2] / "Input" / "SAVE MONEY_20260718.xlsx"
    parser.add_argument("excel", nargs="?", default=str(default_path))
    parser.add_argument("--commit", action="store_true",
                         help="Thuc su xoa (mac dinh chi in ra, khong xoa)")
    args = parser.parse_args()

    path = Path(args.excel)
    if not path.exists():
        print(f"Khong tim thay file: {path}")
        sys.exit(1)
    run(path, commit=args.commit)
