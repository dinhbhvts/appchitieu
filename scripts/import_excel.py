"""One-time importer: load the old Excel workbook into the app database.

What it does
------------
Reads the monthly sheets of the legacy "SAVE MONEY" spreadsheet and turns every
income (THU) / expense (CHI) row into a Transaction row in the app database.
The stock (CK), Logistic and translation sheets are skipped on purpose - stock
data is handled separately.

How to run (from the backend folder, with your virtual environment active)
--------------------------------------------------------------------------
    python -m scripts.import_excel "..\\Input\\SAVE MONEY_20260718.xlsx"

If you leave the path out, it defaults to the file in the Input folder.

Design notes / assumptions (read these - you may want to adjust)
----------------------------------------------------------------
* Person marker: the old GHI CHU column used "H" and "D" to mark who spent the
  money. We map them to the two seeded users below (MARKER_TO_USERNAME). If the
  mapping is backwards for you, just swap the two names and re-run.
* Date: the month and year always come from the sheet name (reliable). The day
  comes from the row's date cell when present; otherwise we reuse the last day
  seen in that sheet, and fall back to day 1. Day precision is not critical
  because all reports group by month/year.
* A row that has BOTH a THU and a CHI value is split into two transactions so
  the money totals stay exactly correct.
* Rows with money but no description get the placeholder content "(khong ghi)".
* The import is idempotent-friendly: pass --fresh to delete existing
  transactions first, so running it twice will not create duplicates.
"""

from __future__ import annotations

import argparse
import datetime
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

# Make "app" importable when this file is run as a module from the backend dir.
sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.core.database import Base, SessionLocal, engine  # noqa: E402
from app.core.seed import seed  # noqa: E402
from app.models.asset import AssetSnapshot  # noqa: E402
from app.models.enums import (  # noqa: E402
    CashFlowType,
    TradeSide,
    TransactionType,
)
from app.models.stock import StockCashFlow, StockTrade  # noqa: E402
from app.models.transaction import Transaction  # noqa: E402
from app.models.user import User  # noqa: E402

# --- Adjust here if needed -------------------------------------------------
# Which app user each old marker belongs to. Swap the names if reversed.
# Confirmed mapping: H = Vo (wife), D = Chong (husband).
# Names must match the seeded user names (Vietnamese with diacritics).
MARKER_TO_USERNAME = {"H": "Vợ", "D": "Chồng"}
# User to assign when the marker is blank or unrecognised.
DEFAULT_USERNAME = "Chồng"
# Sheets that are NOT monthly income/expense sheets - skipped.
SKIP_SHEETS = {"CK", "Logistic", "google translate"}
# Keywords that mark an aggregate / summary row at the bottom of a sheet
# (accents removed, lower-cased). When we meet the first such row in a sheet we
# stop reading that sheet, because everything below is running totals - NOT
# real transactions. This is what keeps the imported figures accurate.
SUMMARY_KEYWORDS = (
    "chot thang", "luy ke", "total", "tong nap", "tong rut", "nap rut",
    "tong thu", "tong chi", "tong cong", "cong thang", "tong ket",
)
# ---------------------------------------------------------------------------


def _strip_accents(text) -> str:
    """Return text without Vietnamese accents, lower-cased, for keyword tests.

    Accepts any type (numbers, None) - they are coerced to string first.
    """
    if text is None:
        return ""
    decomposed = unicodedata.normalize("NFD", str(text))
    no_marks = "".join(c for c in decomposed if unicodedata.category(c) != "Mn")
    return no_marks.lower().strip()


def is_summary_row(content, day_cell) -> bool:
    """True if this row looks like a total/summary line rather than a real one."""
    haystack = _strip_accents(f"{content or ''} {day_cell or ''}")
    return any(kw in haystack for kw in SUMMARY_KEYWORDS)


# Column-header words that must never be treated as an asset name.
ASSET_HEADER_TOKENS = {
    "noi dung", "nap", "mua", "ban", "ma ck", "khoi luong", "gia cp",
    "so tien", "stt", "ngay", "thu", "chi", "du cuoi", "ghi chu", "tong",
    "total", "con lai",
}


def extract_asset_rows(rows: list) -> list[tuple[str, float, str | None]]:
    """Pull the CHOT THANG net-worth block out of one sheet.

    In the asset block, column C (index 2) holds the asset name and column D
    (index 3) holds its value - unlike the transaction area where column C is a
    number. So "C is text AND D is a number" cleanly identifies asset lines. We
    start looking only after the first summary/total row (the block sits at the
    bottom) and stop when we reach the stock (CK) section header.
    """
    # Find where the summary/total block begins.
    start = 0
    for i, raw in enumerate(rows):
        r = list(raw) + [None] * 6
        if is_summary_row(r[1], r[0]):
            start = i
            break

    items: list[tuple[str, float, str | None]] = []
    for i in range(start, len(rows)):
        r = list(rows[i]) + [None] * 6
        name, value, extra = r[2], r[3], r[4]

        # Reached the stock section (NAP ... MUA / MA CK)? Stop.
        if "nap" in _strip_accents(name) and (
            "mua" in _strip_accents(value) or "ma ck" in _strip_accents(extra)
        ):
            break

        if (
            isinstance(name, str)
            and name.strip()
            and _strip_accents(name) not in ASSET_HEADER_TOKENS
            and isinstance(value, (int, float))
        ):
            # Column E is a note ("2 CHI", "gui tiet kiem") only when it is
            # text; when it is a number it is the block's total, which we skip.
            note = extra.strip() if isinstance(extra, str) and extra.strip() else None
            items.append((name.strip(), float(value), note))
    return items


# Words that mark a summary/total row inside the stock section (skip those).
_STOCK_SKIP = ("total", "tong", "loi nhuan", "lai/lo", "lai lo")


def _parse_million_text(text) -> float | None:
    """Parse a Vietnamese short-hand money amount from text, e.g. "rut 4tr7"
    -> 4,700,000; "2,7tr" -> 2,700,000. Returns None if nothing found."""
    t = _strip_accents(text)
    m = re.search(r"(\d+)(?:[.,](\d+))?\s*(?:tr|trieu)(\d*)", t)
    if not m:
        return None
    base = float(m.group(1))
    if m.group(2):        # decimals before "tr": 2,7tr
        base += float("0." + m.group(2))
    if m.group(3):        # digit after "tr": 4tr7
        base += float("0." + m.group(3))
    return base * 1_000_000


def _is_symbol(v) -> bool:
    """True if the value looks like a ticker symbol (2-5 letters)."""
    return isinstance(v, str) and bool(re.fullmatch(r"[A-Z]{2,5}", v.strip().upper()))


def _stock_owner(r: list, note_norm: str) -> str:
    """Work out who a stock row belongs to (returns "Chồng" or "Vợ").

    Priority: an explicit H/D marker (H = Vợ, D = Chồng) in columns N/O/P; then
    the note ("Huế/Huệ" is the wife); otherwise default to Vợ, because this
    brokerage log is overwhelmingly the wife's account.
    """
    for col in (13, 14, 15):
        v = str(r[col]).strip() if r[col] is not None else ""
        if v == "D":
            return "Chồng"
        if v == "H":
            return "Vợ"
    if "hue" in note_norm:
        return "Vợ"
    return "Vợ"


def extract_stock_rows(rows: list, year: int, month: int):
    """Pull the stock section from the bottom of a monthly sheet.

    Columns (fixed by the sub-header "... Mã CK | Khối lượng | Giá CP | ..."):
      A(0) date · B(1) note · C(2) NẠP amount · D(3)/E(4)/F(5) buy sym/qty/price
      · H(7)/I(8)/J(9) sell sym/qty/price.

    Returns (cashflows, trades) as lists of plain dicts (without user_id, which
    the caller fills in).
    """
    # Locate the sub-header row that has "Mã CK" in column D.
    sub = None
    for i, raw in enumerate(rows):
        r = list(raw) + [None] * 20
        if _strip_accents(r[3]) == "ma ck":
            sub = i
            break
    if sub is None:
        return [], []

    cashflows: list[dict] = []
    trades: list[dict] = []
    last_day = 1
    for i in range(sub + 1, len(rows)):
        r = list(rows[i]) + [None] * 20
        note_text = r[1]
        b = _strip_accents(note_text)
        if any(k in b for k in _STOCK_SKIP):
            continue

        day = resolve_day(r[0], year, month, last_day)
        last_day = day
        d = safe_date(year, month, day)

        owner = _stock_owner(r, b)                 # "Chồng" / "Vợ"
        fee = float(r[11]) if isinstance(r[11], (int, float)) else 0.0  # col L

        # --- Cash flow (deposit / withdrawal) ---
        # Deposit amount is in column C (NẠP), withdrawal in column M (RÚT);
        # some withdrawals are only written in the note ("Huế rút 4tr7"). We
        # require the note to say nạp/rút, which also skips the cumulative-total
        # rows (blank note, big running sum).
        c_val, m_val = r[2], r[12]
        if "rut" in b:
            amt = m_val if isinstance(m_val, (int, float)) and abs(m_val) >= 1000 \
                else _parse_million_text(note_text)
            if amt:
                cashflows.append({"date": d, "type": CashFlowType.withdraw,
                                  "amount": abs(float(amt)), "owner": owner,
                                  "note": str(note_text)[:255]})
        elif "nap" in b and isinstance(c_val, (int, float)) and abs(c_val) >= 1000:
            cashflows.append({"date": d, "type": CashFlowType.deposit,
                              "amount": float(c_val), "owner": owner,
                              "note": str(note_text)[:255]})

        # --- Buy order (columns D/E/F, fee in L) ---
        if _is_symbol(r[3]) and isinstance(r[4], (int, float)) and r[4] > 0 \
                and isinstance(r[5], (int, float)) and r[5] > 0:
            trades.append({"date": d, "symbol": r[3].strip().upper(),
                           "side": TradeSide.buy, "quantity": int(r[4]),
                           "price": float(r[5]), "fee": fee, "owner": owner})

        # --- Sell order (columns H/I/J, fee in L) ---
        if _is_symbol(r[7]) and isinstance(r[8], (int, float)) and r[8] > 0 \
                and isinstance(r[9], (int, float)) and r[9] > 0:
            trades.append({"date": d, "symbol": r[7].strip().upper(),
                           "side": TradeSide.sell, "quantity": int(r[8]),
                           "price": float(r[9]), "fee": fee, "owner": owner})

    return cashflows, trades


def parse_month_year(sheet_name: str) -> tuple[int, int] | None:
    """Return (year, month) parsed from a sheet name like 'T7.2026', or None.

    Handles the many naming styles in the file: 'THANG 02.2022', 'Thang 4.2022',
    'T1.2023', etc. We simply look for '<month>.<year>' anywhere in the name.
    """
    m = re.search(r"(\d{1,2})[.\-/ ]+(\d{4})", sheet_name)
    if not m:
        return None
    month, year = int(m.group(1)), int(m.group(2))
    if 1 <= month <= 12 and 2000 <= year <= 2100:
        return year, month
    return None


def parse_amount(value) -> float:
    """Turn a THU/CHI cell into a number. Blank -> 0. Handles text with commas."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Text like "1,200,000" or "1.200.000" or with stray spaces.
    cleaned = re.sub(r"[^\d]", "", str(value))
    return float(cleaned) if cleaned else 0.0


def resolve_day(cell, year: int, month: int, last_day: int) -> int:
    """Work out the day-of-month for a row from its (messy) date cell.

    Returns a day 1..28/31; falls back to last_day (the previous row's day) and
    finally to 1. Never raises.
    """
    if cell is None or cell == "":
        return last_day

    # A real datetime cell: trust its day.
    if isinstance(cell, datetime.datetime):
        return cell.day

    # Numbers: a plain day (1..31), or an Excel serial date if large.
    if isinstance(cell, (int, float)):
        n = float(cell)
        if 1 <= n <= 31:
            return int(n)
        if n > 59:  # looks like an Excel date serial -> convert
            try:
                d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)
                return d.day
            except (OverflowError, ValueError):
                return last_day
        return last_day

    # Text like "2.07", "3/7", "15" -> take the leading number as the day.
    m = re.match(r"\s*(\d{1,2})", str(cell))
    if m:
        day = int(m.group(1))
        if 1 <= day <= 31:
            return day
    return last_day


def safe_date(year: int, month: int, day: int) -> datetime.date:
    """Build a date, clamping the day so e.g. day 31 in February cannot crash."""
    for d in range(day, 0, -1):
        try:
            return datetime.date(year, month, d)
        except ValueError:
            continue
    return datetime.date(year, month, 1)


def import_workbook(path: Path, fresh: bool) -> None:
    """Read the workbook and insert transactions into the database."""
    # Make sure tables exist and the two users / default categories are present.
    Base.metadata.create_all(bind=engine)

    with SessionLocal() as db:
        seed(db)

        # Build a name -> user id lookup so we can assign each row's owner.
        users = {u.name: u.id for u in db.query(User).all()}
        default_uid = users.get(DEFAULT_USERNAME) or next(iter(users.values()))

        if fresh:
            deleted = db.query(Transaction).delete()
            deleted_assets = db.query(AssetSnapshot).delete()
            db.query(StockTrade).delete()
            db.query(StockCashFlow).delete()
            db.commit()
            print(f"Da xoa {deleted} giao dich va {deleted_assets} dong tai "
                  f"san cu + du lieu chung khoan cu (che do --fresh).")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

        stats_by_year: Counter[int] = Counter()
        total_income = 0.0
        total_expense = 0.0
        skipped = 0
        batch: list[Transaction] = []

        # Asset (net-worth) accumulators.
        asset_batch: list[AssetSnapshot] = []
        asset_months = 0

        # Stock accumulators (all assigned to the default user).
        stock_cf_batch: list[StockCashFlow] = []
        stock_trade_batch: list[StockTrade] = []

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            my = parse_month_year(sheet_name)
            if my is None:
                continue
            year, month = my
            ws = wb[sheet_name]

            last_day = 1
            rows = list(ws.iter_rows(values_only=True))

            # --- Asset block (CHOT THANG) for this sheet's month ---
            asset_items = extract_asset_rows(rows)
            if asset_items:
                asset_months += 1
                for name, value, note in asset_items:
                    asset_batch.append(AssetSnapshot(
                        year=year, month=month, name=name, value=value,
                        note=note,
                    ))

            # --- Stock section (nap/rut + mua/ban) for this sheet's month ---
            cfs, trs = extract_stock_rows(rows, year, month)
            for cf in cfs:
                uid = users.get(cf.pop("owner"), default_uid)
                stock_cf_batch.append(StockCashFlow(user_id=uid, **cf))
            for tr in trs:
                uid = users.get(tr.pop("owner"), default_uid)
                stock_trade_batch.append(StockTrade(user_id=uid, **tr))
            # Data starts at row 4 (index 3); rows 1-3 are the header block.
            for raw in rows[3:]:
                # Pad short rows so indexing is always safe.
                r = list(raw) + [None] * (6 - len(raw)) if len(raw) < 6 else raw
                day_cell, content, thu, chi, _ducuoi, ghichu = r[:6]

                # Stop at the first summary/total row - the rest of the sheet is
                # running totals, not transactions.
                if is_summary_row(content, day_cell):
                    break

                income = parse_amount(thu)
                expense = parse_amount(chi)

                # No money at all -> a note/blank row, skip it.
                if income == 0 and expense == 0:
                    skipped += 1
                    continue

                day = resolve_day(day_cell, year, month, last_day)
                last_day = day
                tx_date = safe_date(year, month, day)

                text = (str(content).strip() if content is not None else "")
                if text == "":
                    text = "(khong ghi)"

                marker = str(ghichu).strip() if ghichu is not None else ""
                user_id = users.get(MARKER_TO_USERNAME.get(marker, ""), default_uid)
                # Keep the original marker in the note for traceability.
                note = f"import:{marker}" if marker else "import"

                # One income and/or one expense transaction from this row.
                if income > 0:
                    batch.append(Transaction(
                        date=tx_date, type=TransactionType.income, amount=income,
                        content=text, user_id=user_id, note=note,
                    ))
                    total_income += income
                    stats_by_year[year] += 1
                if expense > 0:
                    batch.append(Transaction(
                        date=tx_date, type=TransactionType.expense, amount=expense,
                        content=text, user_id=user_id, note=note,
                    ))
                    total_expense += expense
                    stats_by_year[year] += 1

        # Insert everything in one commit (fast for thousands of rows).
        db.add_all(batch)
        db.add_all(asset_batch)
        db.add_all(stock_cf_batch)
        db.add_all(stock_trade_batch)
        db.commit()

        print("\n==== KET QUA IMPORT ====")
        print(f"Tong giao dich da tao : {len(batch):,}")
        print(f"Tong THU (income)     : {total_income:,.0f}")
        print(f"Tong CHI (expense)    : {total_expense:,.0f}")
        print(f"So dong bo qua        : {skipped:,}")
        print("So giao dich theo nam :")
        for y in sorted(stats_by_year):
            print(f"   {y}: {stats_by_year[y]:,}")
        print(f"\nTai san (net worth)   : {len(asset_batch):,} dong / "
              f"{asset_months} thang co du lieu")

        n_dep = sum(1 for c in stock_cf_batch if c.type == CashFlowType.deposit)
        n_wd = len(stock_cf_batch) - n_dep
        n_buy = sum(1 for t in stock_trade_batch if t.side == TradeSide.buy)
        n_sell = len(stock_trade_batch) - n_buy
        print(f"Chung khoan           : nap {n_dep}, rut {n_wd}, "
              f"mua {n_buy}, ban {n_sell}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import old Excel into VibeApp DB")
    default_path = Path(__file__).resolve().parents[2] / "Input" / "SAVE MONEY_20260718.xlsx"
    parser.add_argument("excel", nargs="?", default=str(default_path),
                        help="Duong dan toi file Excel")
    parser.add_argument("--fresh", action="store_true",
                        help="Xoa het giao dich cu truoc khi import (tranh trung)")
    args = parser.parse_args()

    path = Path(args.excel)
    if not path.exists():
        print(f"Khong tim thay file: {path}")
        sys.exit(1)
    import_workbook(path, fresh=args.fresh)


if __name__ == "__main__":
    main()
